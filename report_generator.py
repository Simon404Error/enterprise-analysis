# -*- coding: utf-8 -*-
"""
报告生成模块（地区维度）：
  1) 静态备查 Excel：分析结果汇总.xlsx
     - Sheet「地区年度明细」：地区 × 年度，全部指标（比率/增长率 ×100 展示）
     - Sheet「地区汇总」：每个地区一行
     - Sheet「公示分布」：各地区 全部/部分/全部不公示 企业数
  2) 动态文字经营报告（Markdown）：按地区生成，供面板展示/下载。
本项目只做「整个地区」分析，不按单个企业维度输出。
"""
import numpy as np
import pandas as pd

from config import (OUTPUT_XLSX, COL_REGION, COL_YEAR, NUMERIC_COLS, GROWTH_COLS,
                    MAIN_INCOME_WARN, TURNOVER_ACTIVE,
                    PUBLISH_ALL, PUBLISH_PART, PUBLISH_NONE)
from analysis import (region_yearly, add_region_metrics, publish_by_region,
                      publish_by_region_year, region_summary, metrics_for_display, safe_div,
                      data_quality_notes)

# 「地区年度明细」表中删除的列（M-O：是否股权转让、是否对外投资、来源文件）
_DROP_COLS = ['是否股权转让', '是否对外投资', '来源文件']
# 跨年合计（年度=ALL）的标志
_ALL_YEAR = 'ALL'
# 跨年合计行：比率按加总后的数值重算（口径同 add_region_metrics）
_RATIO_MAP = {
    '销售净利率': ('净利润', '销售额或营业收入'),
    '净利润率': ('净利润', '销售额或营业收入'),
    '净资产收益率(ROE)': ('净利润', '所有者权益合计'),
    '资产负债率': ('负债总额', '资产总额'),
    '主营业务收入占比': ('营业总收入中主营业务收入', '销售额或营业收入'),
    '总资产周转率': ('销售额或营业收入', '资产总额'),
}


def _build_all_rows(region_df):
    """每个地区数值列跨年加总，年度记 ALL，比率按合计重算，增长率置空。"""
    g = region_df.groupby(COL_REGION)[NUMERIC_COLS].sum(min_count=1).reset_index()
    g[COL_YEAR] = _ALL_YEAR
    for col, (n, d) in _RATIO_MAP.items():
        g[col] = safe_div(g[n], g[d])
    for c in GROWTH_COLS:
        g[c] = np.nan
    # 标准差沿用地区各年度波动（与逐年明细行一致）
    for col, std_col in [('销售额或营业收入', '销售额标准差'), ('净利润', '净利润标准差')]:
        s = region_df.groupby(COL_REGION)[col].std().rename(std_col)
        g = g.merge(s, left_on=COL_REGION, right_index=True, how='left')
    return g

# 数值列（用于增长趋势的标签/列名对照）
_NUM_LABELS = ['销售额或营业收入', '净利润', '资产总额', '营业总收入中主营业务收入']


# ---------------- 静态 Excel ----------------
def _weighted_main_ratio(detail_df):
    """明细层面「主营业务收入占比」（地区×年度，分子分母同主体）。"""
    if (detail_df is None or detail_df.empty or
            '营业总收入中主营业务收入' not in detail_df.columns or
            '销售额或营业收入' not in detail_df.columns):
        return None
    dd = detail_df[detail_df['销售额或营业收入'] > 0]
    if dd.empty:
        return None
    return (dd.groupby([COL_REGION, COL_YEAR])['营业总收入中主营业务收入'].sum(min_count=1)
            / dd.groupby([COL_REGION, COL_YEAR])['销售额或营业收入'].sum(min_count=1)
            ).rename('主营业务收入占比').reset_index()


def _weighted_main_ratio_pooled(detail_df):
    """地区(跨年合计)层面的同口径主营占比，用于年度=ALL 行。"""
    if (detail_df is None or detail_df.empty or
            '营业总收入中主营业务收入' not in detail_df.columns or
            '销售额或营业收入' not in detail_df.columns):
        return None
    dd = detail_df[detail_df['销售额或营业收入'] > 0]
    if dd.empty:
        return None
    return (dd.groupby(COL_REGION)['营业总收入中主营业务收入'].sum(min_count=1)
            / dd.groupby(COL_REGION)['销售额或营业收入'].sum(min_count=1))


# 全部地区合并时的地区名（明细表末行等）
ALL_REGION = '全部地区'

# 地区"纳税规模"区间预设（下含上不含；上限 None = 不设上限）。阈值单位：元。
# 说明：0-100万 / 100万-1000万 / 1000万以上，按某地区当年合计纳税总额分档。
BRACKET_PRESETS = [
    ('0-100万', 0, 1_000_000),
    ('100万-1000万', 1_000_000, 10_000_000),
    ('1000万以上', 10_000_000, None),
]


def bracket_stats(region_year, years, lo=0.0, hi=None):
    """按地区纳税总额区间逐年度统计：落入该档的地区，其当年
    主营业务收入总额、利润总额，及相邻已选年度的同比增长率。
    region_year：地区×年度指标表（含 纳税总额/营业总收入中主营业务收入/利润总额）。
    lo/hi：纳税区间（元，下含上不含；None 表示不设边界）。
    返回 DataFrame：年度 | 地区数 | 主营业务收入总额 | 主营总额同比增长率 |
                  利润总额 | 利润总额同比增长率(比率)。"""
    lo = 0.0 if lo is None else float(lo)
    hi = float('inf') if hi is None else float(hi)
    years = sorted(int(y) for y in years if not pd.isna(y))
    rows, p_main, p_profit = [], None, None
    for y in years:
        seg = region_year[region_year[COL_YEAR] == y]
        sel = seg[(seg['纳税总额'] >= lo) & (seg['纳税总额'] < hi)]
        main = float(sel['营业总收入中主营业务收入'].sum(min_count=1))
        profit = float(sel['利润总额'].sum(min_count=1))
        gm = ((main - p_main) / p_main) if (p_main is not None and p_main > 0 and main > 0) else np.nan
        gp = ((profit - p_profit) / p_profit) if (p_profit is not None and p_profit > 0 and profit > 0) else np.nan
        rows.append({COL_YEAR: y, '地区数': len(sel), '主营业务收入总额': main,
                     '主营业务收入总额同比增长率': gm, '利润总额': profit,
                     '利润总额同比增长率': gp})
        p_main, p_profit = main, profit
    return pd.DataFrame(rows)


def bracket_table(region_year, presets=BRACKET_PRESETS):
    """把各预设档、各年度拼接为一张长表（供 xlsx 工作表 / 网页）。
    增长率列换算为百分数。列：档位|年度|地区数|主营业务收入总额|主营总额同比增长率|
    利润总额|利润总额同比增长率。"""
    years = sorted(int(y) for y in region_year[COL_YEAR].dropna().unique())
    frames = []
    for label, lo, hi in presets:
        s = bracket_stats(region_year, years, lo, hi)
        s.insert(0, '档位', label)
        frames.append(s)
    out = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if not out.empty:
        out['主营业务收入总额同比增长率'] = (out['主营业务收入总额同比增长率'] * 100).round(2)
        out['利润总额同比增长率'] = (out['利润总额同比增长率'] * 100).round(2)
        out = out[[c for c in ['档位', COL_YEAR, '地区数', '主营业务收入总额',
                               '主营业务收入总额同比增长率', '利润总额', '利润总额同比增长率']
                   if c in out.columns]]
    return out


def _region_block(region_df_single, detail_single):
    """把『单个地区（其 region_df 只含该地区）』拼成「地区年度明细」行：
    逐年度行 + 跨年合计行（年度=ALL），列与明细表一致，含数据提示。"""
    detail = metrics_for_display(region_df_single)
    detail = detail.drop(columns=[c for c in _DROP_COLS if c in detail.columns])
    base_all = _build_all_rows(region_df_single)
    pooled = _weighted_main_ratio_pooled(detail_single)
    if pooled is not None:
        base_all['主营业务收入占比'] = base_all[COL_REGION].map(pooled)
    all_rows = metrics_for_display(base_all)
    all_rows = all_rows.drop(columns=[c for c in _DROP_COLS if c in all_rows.columns])
    notes = data_quality_notes(region_df_single).to_dict()
    blk = pd.concat([detail.sort_values(COL_YEAR), all_rows], ignore_index=True)
    blk['数据提示'] = blk.apply(
        lambda row: notes.get((row[COL_REGION], row[COL_YEAR]), '')
        if row[COL_YEAR] != _ALL_YEAR else '', axis=1)
    return blk


def combined_block(detail_df, label=ALL_REGION):
    """把所有地区当作一个地区合并，返回逐年度行 + 跨年合计(ALL) 行，
    列同「地区年度明细」。用于明细表末行与可视化汇总表。"""
    if detail_df is None or detail_df.empty:
        return pd.DataFrame()
    dd = detail_df.copy()
    dd[COL_REGION] = label
    rc = add_region_metrics(region_yearly(dd), dd)
    return _region_block(rc, dd)


def write_result_excel(detail_df, region_df=None, path=None):
    """生成「分析结果汇总.xlsx」。region_df 为地区×年度指标表。
    明细表按地区分行（逐年度+跨年合计），末行追加「全部地区」合并块。"""
    path = path or OUTPUT_XLSX
    if region_df is None:
        region_df = add_region_metrics(region_yearly(detail_df), detail_df)
    # 主营占比固定改用"分子分母同主体"的企业级口径，规避区内部分企业未报营收致分母失真
    w = _weighted_main_ratio(detail_df)
    if w is not None:
        region_df = region_df.drop(columns=['主营业务收入占比']).merge(
            w, on=[COL_REGION, COL_YEAR], how='left')
        region_df = region_df.sort_values([COL_REGION, COL_YEAR]).reset_index(drop=True)
    # 各地区块（逐年度 + 跨年合计）
    parts = []
    for region in sorted(region_df[COL_REGION].unique()):
        ds = detail_df[detail_df[COL_REGION] == region] if not detail_df.empty else detail_df
        parts.append(_region_block(region_df[region_df[COL_REGION] == region], ds))
    # 末行：全部地区合并块
    cb = combined_block(detail_df)
    if not cb.empty:
        parts.append(cb)
    detail = pd.concat(parts, ignore_index=True)
    summary = metrics_for_display(region_summary(region_df))
    # 公示分布按「地区 × 年度」区分，不计跨年总和
    _, pub_tab = publish_by_region_year(detail_df)

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        detail.to_excel(writer, sheet_name='地区年度明细', index=False)
        summary.to_excel(writer, sheet_name='地区汇总', index=False)
        if not pub_tab.empty:
            pub_tab.to_excel(writer, sheet_name='公示分布', index=False)
        bt = bracket_table(region_df)
        if not bt.empty:
            bt.to_excel(writer, sheet_name='纳税区间统计', index=False)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for cell in ws[1]:
            ws.column_dimensions[cell.column_letter].width = 16
    wb.save(path)
    return path


# ---------------- 文本报告 ----------------
def _pct(v, digits=2):
    """比例 -> '12.34%'；NaN/None -> 'NA'。"""
    if v is None or pd.isna(v):
        return 'NA'
    return f'{v * 100:.{digits}f}%'


def _fmt(v, digits=2):
    if v is None or pd.isna(v):
        return 'NA'
    return f'{v:,.{digits}f}'


def generate_region_report(region_df, region, pub_tab=None):
    """按地区生成经营报告（Markdown）。region_df 为地区×年度指标表。"""
    sub = region_df[region_df[COL_REGION] == region].sort_values(COL_YEAR)
    if sub.empty:
        return f'# 【{region}】\n\n（该地区无数据）'
    years = list(sub[COL_YEAR])
    latest = years[-1]
    cur = sub[sub[COL_YEAR] == latest].iloc[0]

    L = [f'# 【{region}】经营状况报告', '']
    L.append(f'- 覆盖年度：{years[0]} ~ {years[-1]}（共 {len(years)} 个年度）')
    L.append('')

    # ---- 最新年度整体盈利能力 ----
    L.append(f'## 盈利能力（{latest} 年，地区整体）')
    L.append(f'- 销售额（营业收入）：{_fmt(cur["销售额或营业收入"])} 万元')
    L.append(f'- 净利润：{_fmt(cur["净利润"])} 万元')
    L.append(f'- 销售净利率 / 净利润率：{_pct(cur["销售净利率"])}')
    L.append(f'- 净资产收益率(ROE)：{_pct(cur["净资产收益率(ROE)"])}')
    L.append('')

    # ---- 增长趋势（同比） ----
    L.append('## 增长趋势（同比，>1 个年度才有值）')
    for label in _NUM_LABELS:
        col = label + '同比增长率'
        if col not in sub.columns:
            continue
        entries = [f'{int(y)}年 {_pct(v)}' for y, v in zip(sub[COL_YEAR], sub[col])
                   if pd.notna(v)]
        if entries:
            L.append(f'- {label}：' + '、'.join(entries))
    L.append('')

    # ---- 净利润率趋势 → 扩张 / 下降 / N/A ----
    L.append('## 盈利质量与扩张趋势')
    margins = sub['销售净利率']
    if margins.notna().any():
        seq = '、'.join(f'{int(y)}年 {_pct(v)}' for y, v in zip(sub[COL_YEAR], margins))
        L.append(f'- 各年度整体净利润率：{seq}')
        valid = margins.dropna()
        if len(valid) >= 2:
            r = valid.iloc[-1] - valid.iloc[0]
            if r > 1e-9:
                L.append('- 净利润率总体呈上升趋势：反映该地区企业营收扩张、资本扩张、利润增长。')
            elif r < -1e-9:
                L.append('- 净利润率总体呈下降趋势：提示盈利承压。')
            else:
                L.append('- 净利润率总体基本持平。')
        assets = sub['资产总额']
        if len(assets) >= 2:
            L.append(f'- 地区资产总额由 {_fmt(assets.iloc[0])} 万元增至 {_fmt(assets.iloc[-1])} 万元'
                     f'（{int(sub.iloc[0][COL_YEAR])} → {int(sub.iloc[-1][COL_YEAR])}）。')
    else:
        L.append('- 无法计算净利润率（销售额或净利润缺失/为 0），净利润率趋势记为 N/A。')
    L.append('')

    # ---- 主营占比 / 周转率 / 亏损但周转快 ----
    L.append('## 经营与结构提示')
    focus = cur['主营业务收入占比']
    L.append(f'- 主营业务收入占比：{_pct(focus)}'
             + (f'（低于 {MAIN_INCOME_WARN * 100:.0f}%，依赖非主营、经营稳定性可能较差）'
                if pd.notna(focus) and focus < MAIN_INCOME_WARN else ''))
    turnover = cur['总资产周转率']
    L.append(f'- 总资产周转率：{_pct(turnover)}（越高，资产使用效率 / 产能利用率越高）')
    if pd.notna(turnover) and cur['净利润'] < 0 and turnover > TURNOVER_ACTIVE:
        L.append('- 提示：该地区虽亏损但资产周转快，说明运营活跃、产能利用率高。')
    L.append('')

    # ---- 公示分布 ----
    L.append('## 公示状况（本地区内企业）')
    if pub_tab is not None and not pub_tab.empty and (pub_tab[COL_REGION] == region).any():
        row = pub_tab[pub_tab[COL_REGION] == region].iloc[0]
        L.append(f'- 全部公示：{int(row.get(PUBLISH_ALL, 0))} 家；部分公示：{int(row.get(PUBLISH_PART, 0))} 家；'
                 f'全部不公示：{int(row.get(PUBLISH_NONE, 0))} 家。')
    else:
        L.append('- （无企业明细，无法统计公示分布。）')
    L.append('')

    # ---- 异常提示 ----
    L.append('## 异常数据提示')
    tips = []
    if len(years) < 2:
        tips.append('仅有 1 个年度数据，无法计算增长率。')
    n_na = int(sub['销售额同比增长率'].isna().sum() - (1 if len(sub) else 0))
    if n_na > 0:
        tips.append(f'有 {n_na} 条因上年缺失（或断档/上年≤0/由亏转盈）而无有效同比，增长率为 NA（避免无意义%）。')
    if (sub['所有者权益合计'] == 0).any():
        tips.append(f'有 {int((sub["所有者权益合计"] == 0).sum())} 条所有者权益为 0，ROE 为 NA。')
    if (sub['所有者权益合计'] < 0).any():
        tips.append(f'有 {int((sub["所有者权益合计"] < 0).sum())} 条所有者权益为负，ROE 为负值。')
    if not tips:
        tips.append('未发现明显异常。')
    for t in tips:
        L.append(f'- {t}')
    L.append('')
    return '\n'.join(L)


def generate_report(detail_df, region_df=None, region=None):
    """总入口（地区维度）：region 为空时生成全部地区报告合集。"""
    if detail_df is None or detail_df.empty:
        return '# 经营状况报告\n\n（当前筛选条件下无数据）'
    if region_df is None:
        region_df = metrics_for_display(add_region_metrics(region_yearly(detail_df), detail_df))
    _, pub_tab = publish_by_region(detail_df)
    if region:
        return generate_region_report(region_df, region, pub_tab)
    parts = [generate_region_report(region_df, r, pub_tab)
             for r in sorted(region_df[COL_REGION].unique())]
    return '\n\n---\n\n'.join(parts)


if __name__ == '__main__':
    import file_loader
    from analysis import region_yearly, add_region_metrics, metrics_for_display
    d, msgs = file_loader.load_all()
    for m in msgs:
        print(m)
    if not d.empty:
        region = metrics_for_display(add_region_metrics(region_yearly(d), d))
        print(write_result_excel(d, region))
        print(generate_report(d, region)[:400])